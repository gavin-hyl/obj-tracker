from collections import namedtuple

import numpy as np
from matplotlib import pyplot as plt
from openpyxl import load_workbook

from Model import impact_models
from utilities import all_params

plt.style.use('seaborn-v0_8-paper')


def iqr_range(data):
    data.sort()
    Q3 = data[int(len(data) * 3 / 4)]
    Q1 = data[int(len(data) / 4)]
    low = Q1 - 1.5 * (Q3 - Q1)
    high = Q3 + 1.5 * (Q3 - Q1)
    return low, high


def prune_data(sets=None):
    if sets is None:
        sets = [i for i in range(1, 8)]
    wb = load_workbook("Data.xlsx")
    for n in sets:
        params = all_params.get(n)
        sheet = wb[f"{params} (No.{n})"]

        v1x = []
        o1 = []
        v1y = []
        l_vars = []

        row = 2
        while sheet[f"A{row}"].value is not None:
            v1x.append(sheet[f"A{row}"].value)
            v1y.append(sheet[f"B{row}"].value)
            o1.append(sheet[f"C{row}"].value)
            l_vars.append(sheet[f"I{row}"].value)
            row += 1
        row = 2

        v1xL, v1xH = iqr_range(v1x)
        v1yL, v1yH = iqr_range(v1y)
        o1L, o1H = iqr_range(o1)

        for i, o in enumerate(o1):
            if v1yL < v1y[i] < v1yH and v1xL < v1x[i] < v1xH and o1L < o < o1H and l_vars[i] < 2:
                sheet[f"L{row}"].value = 0
            else:
                sheet[f"L{row}"] = 1
            row += 1
        wb.save("Data.xlsx")


def interpolate_sra(x_data, y_data, window=5):
    points = [(x, y_data[i]) for i, x in enumerate(x_data)]
    x_inter = np.linspace(min(x_data), max(x_data), 100)
    y_inter = []
    x_index = 0
    for i in range(len(points) - 1):
        p1 = points[i]
        p2 = points[i + 1]
        while x_inter[x_index] < p1[0]:
            x_index += 1
        while x_index < len(x_inter) and x_inter[x_index] <= p2[0]:
            k = (p2[1] - p1[1]) / (p2[0] - p1[0])
            b = (p2[1] + p1[1] - k * (p2[0] + p1[0])) / 2
            y_inter.append(k * x_inter[x_index] + b)
            x_index += 1
        if x_index >= 100:
            break
    y_inter = np.convolve(y_inter, np.ones(int(window)) / window, 'valid')
    x_out = []
    y_out = []
    for i, _ in enumerate(y_inter):
        x_out.append(x_inter[i + int(window / 2)])
        y_out.append(y_inter[i])
    return x_out, y_out


def prettyplot(ax, loc, x, y_lists, s, labels, colors, y_label, title=r"\\latex", style=None, smooth=False) -> None:
    if style is None:
        style = ["-", "-"]
    for i, y in enumerate(y_lists):
        x_list = x
        # ax[loc[0]][loc[1]].scatter(x, y, s=s, color=colors[i])
        if smooth:
            x_list, y = interpolate_sra(x, y)
        ax[loc[0]][loc[1]].plot(
            x_list, y, label=labels[i], color=colors[i], linestyle=style[i], linewidth=2)
        ax[loc[0]][loc[1]].set_xlabel(r"$\alpha$", fontsize=9)
        ax[loc[0]][loc[1]].set_ylabel(y_label, fontsize=9)
        ax[loc[0]][loc[1]].set_title(title, size=10)
        ax[loc[0]][loc[1]].legend(loc="best")


def main():
    b = load_workbook("Data.xlsx")
    theta_range = (10, 170)
    correction_range = (25, 155)
    datapoint = namedtuple("datapoint", "theta v1x v1y o1 v2x v2y o2 dvx dvy do")
    display = [i + 1 for i in range(7)]
    prune_data(display)
    for exp in display:
        params = all_params.get(exp)
        sheet = b[f"{params} (No.{exp})"]
        exp_points, g_points, test_points = ([] for _ in range(3))
        row = 2
        while sheet[f"D{row}"].value is not None:
            theta = sheet[f"D{row}"].value
            if not (correction_range[0] < theta < correction_range[1]):
                theta -= sheet[f"G{row}"].value / 960
            if theta_range[0] < theta < theta_range[1] and sheet[f"L{row}"].value == 0:
                v1x = sheet[f"A{row}"].value
                v1y = sheet[f"B{row}"].value
                o1 = sheet[f"C{row}"].value
                v2x = sheet[f"E{row}"].value
                v2y = sheet[f"F{row}"].value
                o2 = sheet[f"G{row}"].value
                d = datapoint(theta, v1x, v1y, o1,
                              v2x, v2y, o2, v2x - v1x, v2y - v1y, o2 - o1)
                exp_points.append(d)
            row += 1
        exp_points.sort(key=lambda e: e.theta)
        for _, d in enumerate(exp_points):
            #g, t = impact_models(d.v1x, d.v1y, d.o1, d.v2x, d.v2y, d.o2, d.theta, exp, cdvx=True)
            g, t = impact_models(0, -4, 6.725, d.v2x, d.v2y, d.o2, d.theta, exp, cdvx=True)
            g_points.append(g)
            test_points.append(t)

        size = 20
        sm = True
        thetas = [d.theta for _, d in enumerate(exp_points)]
        experiment_color = "steelblue"
        theory_color = "darkorange"
        _, ax = plt.subplots(2, 2, constrained_layout=True)

        prettyplot(ax=ax, loc=(1, 1), x=thetas, y_lists=[[d.get('de') for _, d in enumerate(test_points)]], s=size,
                   labels=["experiment", "average"], y_label="$E_{ret}$", colors=[experiment_color, experiment_color],
                   style=["-", "--"], title=r"$E_{retained} (\alpha), \%$", smooth=sm)

        prettyplot(ax=ax, loc=(0, 1), x=thetas, y_lists=[[g.dvy for _, g in enumerate(g_points)]], s=size,
                   labels=["model", "model"], y_label="$\Delta v_y$", colors=[experiment_color, theory_color],
                   title=r"$\Delta v_y (\alpha), m/s$", smooth=sm)

        prettyplot(ax=ax, loc=(1, 0), x=thetas, y_lists=[[g.do/360 for _, g in enumerate(g_points)]], s=size,
                   labels=["model", "model"], y_label="$\Delta \omega$", colors=[experiment_color, theory_color],
                   title=r"$\Delta \omega (\alpha), rev/s$", smooth=sm)

        prettyplot(ax=ax, loc=(0, 0), x=thetas, y_lists=[[g.dvx for _, g in enumerate(g_points)]], s=size,
                   labels=["model", "model"], y_label="$\Delta v_x$", colors=[experiment_color, theory_color],
                   title=r"$\Delta v_x (\alpha), m/s$", smooth=sm)

        plt.suptitle(t=f"{params} (No.{exp})", size=12)
        plt.show()
        # plt.savefig(f"Results/exp{exp}.jpg", dpi=1080)


if __name__ == "__main__":
    main()

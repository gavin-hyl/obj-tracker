import string
from collections import namedtuple

import cv2 as cv
import numpy as np
# import this (sorry I had to)
from matplotlib import pyplot as plt
from openpyxl import load_workbook

from utilities import all_params

FPS = 1920
T_FRAME = 1 / FPS
WHITE = (255, 255, 255)
BLUE = (255, 0, 0)
GREEN = (0, 255, 0)
RED = (0, 0, 255)
BLACK = (0, 0, 0)
WIDTH = 2e-2

frame_data = namedtuple('frame_data', ['num', 'x', 'y', 'angle', 'l', 'w'])
vid_data = namedtuple('vid_data', ['v1x', 'v1y', 'o1', 'theta', 'v2x', 'v2y', 'o2', 'lv1', 'lv2', 'wv1', 'wv2'])
vid_frame_data = []


def process_video(video_name, start, vid_len, crop_top, crop_bottom,
                  crop_left, crop_right, binary_thresh, flip, show, write, excel_worksheet, plot, buffer_frames):
    # Preparing RGB frames
    vid = cv.VideoCapture(video_name)
    frames = []
    success, img = vid.read()
    scale_percent = 50
    width = int(img.shape[1] * scale_percent / 100)
    height = int(img.shape[0] * scale_percent / 100)
    dim = (width, height)
    while success:
        img = cv.resize(img, dim)
        if flip:
            img = cv.flip(img, -1)
        img = img[crop_top:dim[1] - crop_bottom, crop_left:dim[0] - crop_right]
        frames.append(img)
        success, img = vid.read()

    # Process frames in specified range.
    # This is a separate loop since operations like rectangle fitting are computation-heavy, so
    # we don't want them for every frame
    for i in range(start, start + vid_len):
        # blur, to binary, create rgb dummy
        cropped = cv.GaussianBlur(frames[i], (5, 5), 1)       
        gray = cv.cvtColor(cropped, cv.COLOR_BGR2GRAY)
        binary_frame = cv.threshold(gray, binary_thresh, 255, cv.THRESH_BINARY)[1]
        binary_canvas = cv.cvtColor(binary_frame, cv.COLOR_GRAY2RGB)

        # convex, find the largest contour
        contours, _ = cv.findContours(binary_frame, cv.RETR_TREE, cv.CHAIN_APPROX_NONE)
        hulls = [cnt for cnt in contours]
        hulls.sort(key=lambda hull: cv.contourArea(hull))
        target = hulls[-1]

        # find, store, and draw the bounding box
        rect = cv.minAreaRect(target)
        center = rect[0]
        lw = rect[1]
        if lw[1] > lw[0]:
            length = lw[1]
            width = lw[0]
        else:
            length = lw[0]
            width = lw[1]
        box = np.int32(cv.boxPoints(rect))
        cv.drawContours(binary_canvas, [box], 0, color=GREEN, thickness=2)

        # find, store, and draw the ellipse
        ellipse_center, maj_min_axes, angle = cv.fitEllipse(target)
        position = np.int32(ellipse_center)
        binary_canvas = cv.ellipse(binary_canvas, np.int32(ellipse_center), np.int32(maj_min_axes),
                                   np.int32(angle), 0, 360, RED, thickness=2)
        cv.circle(binary_canvas, np.int32(position), 2, BLACK, thickness=-1)

        # storing everything
        angle = 270 - angle if angle >= 90 else (90 - angle)
        vid_frame_data.append(frame_data(i, center[0], center[1], angle, length, width))

        # display fitting results
        if show:
            cv.putText(cropped, f'frame{i}, angle{int(angle)}', org=(10, 100),
                       fontFace=cv.FONT_HERSHEY_SIMPLEX, fontScale=1, color=WHITE, thickness=1, lineType=2)
            cv.imshow('fittings', binary_canvas)
            cv.imshow('rgb', cropped)
            cv.waitKey(0)
            if cv.waitKey(3) & 0xFF == ord('q'):
                break

    # scale & speeds
    scale = np.mean([f.w for f in vid_frame_data]) / WIDTH
    vx_t = np.diff([f.x * FPS / scale for f in vid_frame_data])
    vy_t = np.diff([f.y * FPS / scale for f in vid_frame_data])
    a_t = [f.angle for f in vid_frame_data]
    o_t = []
    for i in range(len(a_t) - 1):
        a1 = a_t[i]
        a2 = a_t[i + 1]
        if a2 - a1 < -100:  # only occurs when the angle goes from 180->0 (pos spin)
            d_a = 180 + a2 - a1
        elif a2 - a1 > 100:  # only occurs when the angle goes from 0->180 (neg spin)
            d_a = -(180 - (a2 - a1))
        else:
            d_a = a2 - a1
        o_t.append(d_a * 1920)

    # impact frame related
    vid_frame_data.sort(key=lambda e: e.y, reverse=True)
    impact_frame = int(np.mean([vid_frame_data[i].num for i in range(3)]))
    before = impact_frame - buffer_frames - start
    after = impact_frame + buffer_frames - start

    # separate speeds
    v1x, v2x = np.mean(vx_t[:before]), np.mean(vx_t[after:])
    v1y, v2y = (np.mean(vy_t[:before][-20:]) - 0.05), (np.mean(vy_t[after:][:20]) + 0.05)
    o1, o2 = np.mean(o_t[:before]), np.mean(o_t[after:])
    angle = vid_frame_data[impact_frame].angle
    t = [i / FPS for i, _ in enumerate(vx_t)]
    t1, t2 = t[:before], t[after:]

    l_va1, l_va2 = np.var([f.l for f in vid_frame_data[:before]]), np.var([f.l for f in vid_frame_data[after:]])
    w_va1, w_va2 = np.var([f.w for f in vid_frame_data[:before]]), np.var([f.w for f in vid_frame_data[after:]])

    if plot:
        _, axes = plt.subplots(1, 3, constrained_layout=True)
        axes[0].plot(t1, vx_t[:before], label='bef. impact')
        axes[0].plot(t2, vx_t[after:], label='bef. impact')
        axes[0].plot(t1, [v1x] * len(t1), label='av. bef. impact')
        axes[0].plot(t2, [v2x] * len(t2), label='av. aft. impact')
        axes[0].legend(loc='best')
        axes[0].set_title('vx, m/s', size=18)
        axes[1].plot(t1, vy_t[:before], label='bef. impact')
        axes[1].plot(t2, vy_t[after:], label='aft. impact')
        axes[1].scatter(max(t1), v1y, s=40, marker='X', color='black', label='av. bef. impact')
        axes[1].scatter(min(t2), v2y, s=40, marker='X', color='black', label='av. aft. impact')
        axes[1].legend(loc='best')
        axes[1].set_title('vy, m/s', size=18)
        axes[2].plot(t1, o_t[:before], label='bef. impact')
        axes[2].plot(t2, o_t[after:], label='aft. impact')
        axes[2].plot(t1, [o1] * len(t1), label='av. bef. impact')
        axes[2].plot(t2, [o2] * len(t2), label='av. aft. impact')
        axes[2].legend(loc='best')
        axes[2].set_title('omega, deg/s', size=18)
        plt.show()

    if write:
        wb = load_workbook('Data.xlsx')
        ws = wb[excel_worksheet]
        row = 1
        while not ws[f'A{row}'].value is None:
            row += 1
        for i, v in enumerate([v1x, v1y, o1, angle, v2x, v2y, o2, l_va1, l_va2, w_va1, w_va2]):
            letter = string.ascii_uppercase[i]
            ws[f'{letter}{row}'] = v
        wb.save('Data.xlsx')
        print(f'{video_name} write successful')


def main():
    i, exp, stop = 1, 1, 0
    params = all_params.get(exp)
    while True:
        try:
            vid_path = f'{params}/{i}'
            process_video(video_name=f'Videos/{vid_path}.mp4', excel_worksheet=f'{params} (No. {exp})',
                          flip=True,
                          show=True,
                          plot=True,
                          write=False,
                          crop_top=0,
                          crop_bottom=30,
                          crop_left=0,
                          crop_right=0,
                          start=90,
                          vid_len=225,
                          buffer_frames=5,
                          binary_thresh=130,)
        except IndexError:
            stop += 1
            if stop >= 10:
                break
        i += 1


if __name__ == '__main__':
    main()
